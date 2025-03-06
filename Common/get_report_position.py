from openpyxl import load_workbook
from openpyxl.styles import Font
from Common.config import Config


class GetReportPosition:
    def __init__(self, file_path, sheet_name):
        self.file_path = file_path
        self.sheet_name = sheet_name

    def find_scenario_position_by_keyword(self, keyword):
        wb = load_workbook(self.file_path)
        sheet = wb[self.sheet_name]
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and keyword in cell.value:
                    # 获取单元格的行和列（注意行和列是从1开始的）, 返回x, y
                    wb.close()
                    return [cell.row, cell.column]

    def get_ae_stability_light_lux_key_word_position(self, key_word):
        key_position = self.find_scenario_position_by_keyword(key_word)
        return key_position

    def get_all_ae_stability_light_lux_position(self, key_word):
        # 关键字“光照亮度”占了两行，所以函数需要从 +2 开始
        positions = {}
        key_position = self.get_ae_stability_light_lux_key_word_position(key_word)
        positions["lux_1000"] = [key_position[0] + 2, key_position[1] + 1]
        positions["lux_500"] = [key_position[0] + 3, key_position[1] + 1]
        positions["lux_250"] = [key_position[0] + 4, key_position[1] + 1]
        positions["lux_128"] = [key_position[0] + 5, key_position[1] + 1]
        positions["lux_64"] = [key_position[0] + 6, key_position[1] + 1]
        positions["lux_32"] = [key_position[0] + 7, key_position[1] + 1]
        positions["lux_16"] = [key_position[0] + 8, key_position[1] + 1]
        positions["lux_8"] = [key_position[0] + 9, key_position[1] + 1]
        return positions

    def get_ae_convergence_positions(self, keyword, data_length):
        position = []
        key_position = self.find_scenario_position_by_keyword(keyword)
        for i in range(1, data_length + 1):
            position.append([key_position[0] + i, key_position[1]])
        return position

    def get_results_statistics_position(self, data_dict):
        # data_dict{"50lux":[1, 2], "400lux":[2, 3], "1000":[5,4]}
        x_lux_positions = [data_dict["lux_50"][0], data_dict["lux_400"][0], data_dict["lux_1000"][0]]
        x_max = max(x_lux_positions)
        y_max = data_dict["lux_1000"][1]

        lx_50_result_position = [x_max + 1, y_max - 2]
        lx_400_result_position = [x_max + 1, y_max - 1]
        lx_1000_result_position = [x_max + 1, y_max]

        frame_que_position = [x_max + 1, y_max - 3]
        # 最后一行的数据
        result_dict = {"calculate_result": frame_que_position, "lx_50_frames_sum": lx_50_result_position, "lx_400_frames_sum": lx_400_result_position, "lx_1000_frames_sum": lx_1000_result_position}
        return result_dict

    def get_frame_num_position(self, keyword, max_list_len):
        positions = []
        # 关键字为 50lx
        key_position = self.find_scenario_position_by_keyword(keyword)
        frame_num_position = [key_position[0], key_position[1] - 1]
        for i in range(1, max_list_len + 1):
            positions.append([frame_num_position[0] + i, frame_num_position[1]])
        return positions


if __name__ == '__main__':
    g_p_s = GetReportPosition(Config.original_template_path)
    print(g_p_s.get_all_ae_stability_light_lux_position(Config.light_lux, Config.ae_stability_sheet_name))
