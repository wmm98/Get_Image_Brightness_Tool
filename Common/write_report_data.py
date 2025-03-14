from openpyxl import load_workbook
from Common.config import Config
from Common.get_report_position import GetReportPosition
from openpyxl.styles import Alignment, Border, Side, PatternFill
from openpyxl.chart import LineChart, Reference


class WriteReport:
    def __init__(self, template_path, sheet_name):
        self.sheet_name = sheet_name
        self.template_path = template_path
        self.report_position = GetReportPosition(self.template_path, self.sheet_name)

    def get_border(self):
        # 创建一个细边框样式
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))
        return thin_border

    def write_ae_stability_data(self, ae_position, ae_value):
        try:
            wb = load_workbook(self.template_path)
            sheet = wb[self.sheet_name]
            for k in ae_position:
                cell = sheet.cell(row=ae_position[k][0], column=ae_position[k][1])
                cell.value = ae_value[k]
            wb.save(self.template_path)
        finally:
            wb.close()

    def write_ae_convergence_data(self, ae_position, ae_value):
        try:
            wb = load_workbook(self.template_path)
            sheet = wb[self.sheet_name]
            for k in range(len(ae_position)):
                cell = sheet.cell(row=ae_position[k][0], column=ae_position[k][1])
                cell.value = ae_value[k]
                # cell.border = self.get_border()
                cell.alignment = Alignment(horizontal='center', vertical='center')
            wb.save(self.template_path)
        finally:
            wb.close()

    def write_ae_convergence_number_data(self, num_positions):
        try:
            wb = load_workbook(self.template_path)
            sheet = wb[self.sheet_name]
            num = 1
            for n in num_positions:
                cell = sheet.cell(row=n[0], column=n[1])
                cell.value = num
                # cell.border = self.get_border()
                cell.alignment = Alignment(horizontal='center', vertical='center')
                num += 1
            wb.save(self.template_path)
        finally:
            wb.close()

    def write_result_data(self, result_positions, result_values):
        try:
            wb = load_workbook(self.template_path)
            sheet = wb[self.sheet_name]
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            for result in result_positions:
                cell = sheet.cell(row=result_positions[result][0], column=result_positions[result][1])
                if result == "calculate_result":
                    cell.value = "%s" % str(result_values[result])
                else:
                    cell.value = "%s帧" % str(result_values[result])
                cell.fill = yellow_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
            wb.save(self.template_path)
        finally:
            wb.close()

    def write_border(self, x_min, x_max, y_min, y_max):
        try:
            wb = load_workbook(self.template_path)
            sheet = wb[self.sheet_name]
            for x in range(x_min, x_max + 1):
                for y in range(y_min, y_max + 1):
                    # print([x, y])
                    cell = sheet.cell(row=x, column=y)
                    cell.border = self.get_border()
                    if x == x_max:
                        sheet.column_dimensions[cell.column_letter].width = 14
                        sheet.row_dimensions[cell.row].height = 20
                    else:
                        sheet.column_dimensions[cell.column_letter].width = 14
                        sheet.row_dimensions[cell.row].height = 15

            wb.save(self.template_path)
        finally:
            wb.close()

    def writ_line_chart(self, param_dict):
        # param_dict {"50lux":[[1, 2], len(data)], "400lux":[[1, 2], len(data)], "1000lux":[[1, 2], len(data)]}
        wb = load_workbook(self.template_path)
        sheet = wb[self.sheet_name]

        chart = LineChart()
        chart.title = "AE收敛速度"
        chart.x_axis.title = '帧数'
        chart.y_axis.title = '亮度均值'
        # 画图表需要 min_row：row-1   row为横坐标
        data_50lux = Reference(sheet, min_col=param_dict["50lux"][0][1], min_row=param_dict["50lux"][0][0] - 1, max_col=param_dict["50lux"][0][1], max_row=param_dict["50lux"][1] + 3)
        data_400lux = Reference(sheet, min_col=param_dict["400lux"][0][1], min_row=param_dict["400lux"][0][0] - 1, max_col=param_dict["400lux"][0][1], max_row=param_dict["400lux"][1] + 3)
        data_1000lux = Reference(sheet, min_col=param_dict["1000lux"][0][1], min_row=param_dict["1000lux"][0][0] - 1, max_col=param_dict["1000lux"][0][1], max_row=param_dict["1000lux"][1] + 3)

        chart.add_data(data_50lux, titles_from_data=True)
        chart.add_data(data_400lux, titles_from_data=True)
        chart.add_data(data_1000lux, titles_from_data=True)

        chart.width = 25
        chart.height = 10

        sheet.add_chart(chart, "F3")

        wb.save(self.template_path)
        wb.close()



if __name__ == '__main__':
    w_r = WriteReport(Config.template_path, Config.sheet_name)
    w_r.write_scenario_data(Config.f_data_path, Config.r_F_light)
    w_r.write_hj_data()

    # w_r = WriteReport(Config.template_path, Config.sheet_name)
    # report_position = GetReportPosition(Config.template_path, Config.sheet_name)

    # 灰阶测试
    # csv = CSVTestData(Config.hj_data_path)
    # hj_data = csv.get_hj_relate_data(csv.read_csv_to_matrix())
    # print(hj_data)

